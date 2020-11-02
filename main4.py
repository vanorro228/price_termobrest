import csv
import math
from collections import Counter, OrderedDict
import openpyxl
import pandas as pd
from selenium import webdriver
import re
import requests
import bs4
from collections import OrderedDict
import urllib.request
from lxml import html
from selenium.common.exceptions import StaleElementReferenceException
from collections import OrderedDict
from urllib.request import urlretrieve

"""w = []
w1 = []
f1 = []
wb1 = openpyxl.load_workbook('55448060.xlsx')
ok = wb1.get_sheet_by_name("Тексты")
for i in range(12, 3065):
    w.append(ok.cell(row=i, column=6).value)
for i in OrderedDict(Counter(w)).values():
    f1.append(i)
wb1.close()
wb2 = openpyxl.load_workbook('test14882.xlsx')
ok1 = wb2.get_sheet_by_name("оплата")
for i in range(1, 194):
    w1.append(ok1.cell(row=i, column=1).value)


def cou(k):
    return f1[k]


def oks():
    wb = xlwt.Workbook()
    ws = wb.add_sheet("оплата", cell_overwrite_ok=True)
    k = -1
    m = 0
    for line in w1:
        for j in range(5):
            k += 1
            ws.write(k, 0, line)
    wb.save("auff.xls")

def csv_dict_reader(f):
    reader = csv.DictReader(f, delimiter=';')
    d = []
    wb = xlwt.Workbook()
    ws = wb.add_sheet("оплата", cell_overwrite_ok=True)
    k = -1
    m = 0
    for line in reader:
        try:
            m += 1
            if line["product_price"] != '':
                line["product_price"] = str(line["product_price"][:-3])
                print(m, f1[m])
                if line["product_price"] != '0':
                    if int(line["product_price"]) % 50 == 0:
                        for j in range(5):
                            k += 1
                            ws.write(k, 0, "Цена: " + str(round(int(line["product_price"]), -2)+100) + " руб.")
                    else:
                        for j in range(5):
                            k += 1
                            ws.write(k, 0, "Цена: " + str(round(int(line["product_price"]),-2)) + " руб.")
                else:
                    for j in range(5):
                        k += 1
                        ws.write(k, 0, "Узнать цену")
        except TypeError:
            ws.write(k, 0, "АУУУУУУУУУУУУУУУУУФФФФФФФФ ПАМАГИТЕ!!!")
        except ValueError:
            ws.write(k, 0, "АУУУУУУУУУУУУУУУУУФФФФФФФФ ПАМАГИТЕ!!!")
    wb.save("tes9.xls")"""

browser = webdriver.Chrome(executable_path="C:\\Users\\user\\PycharmProjects\\pythonProject1\\chromedriver.exe")

eq = ['Датчики DUNGS',
      'reguljatory-dungs',
      'Клапаны DUNGS',
      'mul-tibloki-dungs',
      'bloki-kontrolja-germetichnosti-dungs',
      'Принадлежности',
      'fil-try-dungs',
      'Приводы DUNGS',
      'Заслонки DUNGS',
      'Краны DUNGS',
      'Автоматика DUNGS',
      'Запчасти DUNGS']


def transliterate(name):
    slovar = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'yo',
              'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'i', 'к': 'k', 'л': 'l', 'м': 'm', 'н': 'n',
              'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'yu', 'ф': 'f', 'х': 'h',
              'ц': 'c', 'ч': 'ch', 'ш': 'sh', 'щ': 'sht', 'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e',
              'ю': 'u', 'я': 'ya', 'А': 'a', 'Б': 'b', 'В': 'v', 'Г': 'g', 'Д': 'd', 'Е': 'e', 'Ё': 'yo',
              'Ж': 'zh', 'З': 'z', 'И': 'i', 'Й': 'y', 'К': 'k', 'Л': 'l', 'М': 'm', 'Н': 'n',
              'О': 'o', 'П': 'p', 'Р': 'r', 'С': 's', 'Т': 't', 'У': 'u', 'Ф': 'f', 'Х': 'h',
              'Ц': 'c', 'Ч': 'ch', 'Ш': 'sh', 'Щ': 'sht', 'Ъ': '', 'Ы': 'y', 'Ь': '', 'Э': 'E',
              'Ю': 'yu', 'Я': 'ya', 'A': 'а', 'B': 'b', 'V': 'v', 'G': 'g', 'D': 'd', 'E': 'e', 'Z': 'z', 'I': 'i',
              'O': 'o', 'P': 'p', 'R': 'r', 'S': 's', 'T': 't', 'U': 'u', 'F': 'f', 'H': 'h', 'W': 'w', '" ': '-',
              'C': 'c', 'K': 'k', 'L': 'l', 'M': 'm', ',': '-', '?': '', ' ': '-', 'ґ': '', 'ї': '', ')': '',
              'Ґ': 'g', 'Ї': 'i', 'Є': 'e', 'N': 'n', ' - ': '-', '/': '-', '(': '', '.': '-', '"': ''}

    # Циклически заменяем все буквы в строке
    for key in slovar:
        name = name.replace(key, slovar[key])
    name = re.sub('-{2,}', '-', name)
    return name


def csv_writer(data):
    with open('category.csv', "w", newline='', encoding='utf-8') as out_file:
        writer = csv.writer(out_file)
        for row in data:
            writer.writerow([row])


def csv_writer1(data):
    with open('short_discr.csv', "w", newline='', encoding='utf-8') as out_file:
        writer = csv.writer(out_file)
        for row in data:
            writer.writerow([row])


def output():
    a = []
    ref = ['https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-3-a5.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-10-a5.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-10-a5-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-50-a5.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-50-a5-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-150-a5.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-150-a5-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-500-a5.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a5/gw-500-a5-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-3-a6.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-10-a6.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-50-a6.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-150-a6.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-500-a6.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-50-a6-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-150-a6-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a6/gw-500-a6-1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2/lgw-3-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2/lgw-10-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2/lgw-50-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2/lgw-150-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4/lgw-3-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4/lgw-10-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4/lgw-50-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4/lgw-150-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2p/lgw-3-a2p.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2p/lgw-10-a2p.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2p/lgw-50-a2p.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2p/lgw-150-a2p.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2-7/lgw-1-5-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2-7/lgw-3-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2-7/lgw-6-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2-7/lgw-10-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a2-7/lgw-30-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-a2-7/ks-150-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-a2-7/ks-300-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-a2-7/ks-600-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-a2-7/ks-1000-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-a2-7/ks-3000-a2-7.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/ub-50-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/ub-150-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/ub-500-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/nb-50-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/nb-150-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a4/nb-500-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/ub-50-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/ub-150-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/ub-500-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/nb-50-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/nb-150-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ub-nb-a2/nb-500-a2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-hp/gw-500-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-hp/gw-500-a4-hp-ip54m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-hp/gw-2000-a4-hp-ip54m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-hp/gw-6000-a4-hp-ip54m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-hp/gw-2000-a4-hp-g3.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-2-hp/gw-500-a4-2-hp-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-2-hp/gw-2000-a4-2-hp-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-2-hp/gw-6000-a4-2-hp-p65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-gw-a4-2-hp/gw-6000-a4-2-hp-sgs.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a1/lgw-1-5-a1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a1/lgw-3-a1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a1/lgw-10-a1.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4/ggw-3-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4/ggw-10-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4/ggw-50-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4/ggw-150-a4.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4-2/ggw-3-a4-2-ip65m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4-2/ggw-10-a4-2-ip65m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4-2/ggw-50-a4-2-ip65m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ggw-a4-2/ggw-150-a4-2-ip65m.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-3-a4-2-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-10-a4-2-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-50-a4-2-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-150-a4-2-ip65.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-3-a4-2-g3.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-a4-2/lgw-10-a4-2-g3.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-1-5-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-3-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-5-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-6-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-10-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-lgw-c2/lgw-30-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-150-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-300-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-300-1-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-500-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-600-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-1000-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/dungs-ks-c2/ks-3000-c2.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/shteker-dungs-210318.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-sborki-ldjynogo-rele-dungs-213910.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nippel-privarnoy-dungs-1203402.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/ugol-krepezhny-dungs-230274.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/perehodnik-dungs-1203406.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/perehodnik-dungs-1500395.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/montazhnaya-plata-dungs-230302.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/shtucer-dungs-230306.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/shlang-dungs-1203405.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-montazha-lamp-231772.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-montazha-lamp-231773.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-montazha-lamp-231774.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-montazha-lamp-248239.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/nabor-dlya-montazha-lamp-248240.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/oporniy-ugolok-metallicheskiy-dungs-230288.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/komplekt-dlya-montazha-dungs-242771.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/komplekt-dlya-montazha-dungs-223280.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/izmeritelnyi-patrubok-dungs190808115350.html',
           'https://dungs.rusmark.ru/datchiki-dungs/prinadlezhnosti/izmeritelnyi-patrubok-dungs-1-8-230397190807134508.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-503.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-505.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-507.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-510.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-515.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-rp-3-8-rp-2/frs-520.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5040.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5050.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5065.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5080.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5100.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5125.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/dungs-frs-dn-40-dn-150/frs-5150.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-503.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-505.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-507.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-510.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-515.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-rp-3-8-rp-2/frng-520.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5040.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5050.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5065.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5080.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5100.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5125.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frng-dn-40-dn-150/frng-5150.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-515.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-520.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-5040.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-5050.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-5065.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-5080.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frn/frn-5100.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-fri/fri-705-5-230472.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-fri/fri-707-6-230473190807110217.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-fri/fri-710-6-230474.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-fri/fri-712-6-230475.html',
           'https://dungs.rusmark.ru/reguljatory-dungs/tip-frsbv/frsbv-1010.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-503-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-507-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-512-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-520-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5040-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5050-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5065-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5080-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5100-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5125-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-525-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5080-11190926132149.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-d/dmv-d-5065-11-221691.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-503-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-507-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-512-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-520-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5040-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5050-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5065-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5080-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5100-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5125-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-525-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-dle/dmv-dle-5065-11-221692.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2200.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-203-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-205-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-207-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-210-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-215-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-225-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-503-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-505-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-507-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-515-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2040-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2050-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2080-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2100-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2125-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-2150-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-5040-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-5065-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvd/mvd-5100-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-203-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-205-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-207-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-210-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-215-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-220-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-225-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-2040-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-2050-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-2065-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-2080-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-2100-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-503-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-507-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-515-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-520-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mvdle/mvdle-5050-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-525-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-5065-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-5080-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-5100-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-d-5065-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-d-5080-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-d-5100-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-d-5125-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-dle-5065-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-dle-5080-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-dmv-12/dmv-dle-5100-12.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-lgv-507-5/lgv-507-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-407-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-410-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-415-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-420-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-4040-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrdle/zrdle-4050-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mv/mv-502-0.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mv/mv-502-1.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-mv/mv-502-1-1.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-11-eco/dmv-525-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-11-eco/dmv-5065-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-11-eco/dmv-5080-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-11-eco/dmv-5100-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-11-eco/dmv-5125-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-d-11-eco/dmv-d-525-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-d-11-eco/dmv-d-5065-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-d-11-eco/dmv-d-5080-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-d-11-eco/dmv-d-5100-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-d-11-eco/dmv-d-5125-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-dle-11-eco/dmv-dle-5065-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-dle-11-eco/dmv-dle-525-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-dle-11-eco/dmv-dle-5080-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-dle-11-eco/dmv-dle-5100-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-dle-11-eco/dmv-dle-5125-11-eco.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrd/zrd-407-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrd/zrd-410-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrd/zrd-415-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrd/zrd-420-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrle/zrle-407-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrle/zrle-410-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrle/zrle-415-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-zrle/zrle-420-5.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-505.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-507.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-510.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-512.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-515.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv/sv-520.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-505.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-507.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-510.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-512.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-515.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-d/sv-d-520.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-505.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-507.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-510.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-512.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-515.html',
           'https://dungs.rusmark.ru/klapany-dungs/dungs-sv-dle/sv-dle-520.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5065-11-s82.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5080-11-s82.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5080-11-s302.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-507-11-s300.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-512-11-s80.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-512-11-s300.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-520-11-s20.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-520-11-s80.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-520-11-s300.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5065-11-s20.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5065-11-s80.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5080-11-s20.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5080-11-s300.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5100-11-s20.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5100-11-s80.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-5100-11-s300.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-507-11-s82.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-507-11-s302.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-512-11-s302.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-se/dmv-se-520-11-s22.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-507-11-s12.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-512-11-s12.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-520-11-s12.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-507-11-s32.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-512-11-s32.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-520-11-s32.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-5065-11-s12.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-5065-11-s30.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-5080-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-5100-11.html',
           'https://dungs.rusmark.ru/klapany-dungs/tip-dmv-vef/dmv-vef-5125-11.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-403-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-407-b01-s20-226561.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-412-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-415-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-405-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-410-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-412-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-415-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-405-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-410-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-420-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-407-b01-s22-225015.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-420-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-405-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-407-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-410-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-412-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-407-b01-s22-253535.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-407-b01-s20-231165.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01s20-22/mb-dle-412-b01-s20200605130043.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-403-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-407-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-415-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-410-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-412-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-415-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-420-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-405-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-410-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-412-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-420-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-407-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-dle-b01-s50-52/mb-dle-405-b01-s50-229778.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-405-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-407-b01-s20-226553.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-410-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-412-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-415-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-420-b01-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-407-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-410-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-412-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-415-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-420-b01-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-405-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-407-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-410-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-412-b07-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-407-b01-s20-231170.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s20-22/mb-zrdle-412-b01-s20-231174.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-407-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-410-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-412-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-415-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-420-b01-s50.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-407-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-410-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-412-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-415-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-420-b01-s52.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-zrdle-b01-s50-52/mb-zrdle-407-b01-s50-242563.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-407-b01-s10.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-412-b01-s10.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-412-b01-s30.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-415-b01-s30.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-420-b01-s10.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-420-b01-s30.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-425-b01-s10.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-407-b01-s30.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-415-b01-s10.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-425-b01-s30.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-412-b01-s10-232473.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-407-b01-s30-242564.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-412-b01-s30191218142626.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-420-b01-s10-228037.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-420-b01-s10-230628.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-407-b01-s10-238809.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s10-30/mb-vef-407-b01-s10-247543.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-407-b01-s12.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-407-b01-s32-241117.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-412-b01-s12.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-412-b01-s32.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-415-b01-s32.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-420-b01-s12.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-420-b01-s32.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-425-b01-s12.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-425-b01-s32.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-415-b01-s12.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mb-vef-s12-32/mb-vef-420-b01-s32-258850.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-300-se-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-700-se-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-1200-se-s22.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-300-se-s82.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-700-se-s82.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-1200-se-s82.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-300-se-s302.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-700-se-s302.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-1200-se-s302.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-1900-se-65.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-3100-se-80.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-5000-se-100.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-300-se-s02.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-700-se-s02.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-1200-se-s02.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-se/mbc-300-se-s22-250282.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-300-vef.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-700-vef.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-1200-vef.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-1900-vef-65.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-3100-vef-80.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-vef/mbc-5000-vef-100.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-dle/mbc-65-dle-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-dle/mbc-65-dle-s40.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-dle/mbc-120-dle-s20.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-dle/mbc-120-dle-s40.html',
           'https://dungs.rusmark.ru/mul-tibloki-dungs/dungs-mbc-dle/mbc-65-dle-s20190924164347.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s01.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s02-0.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s02-1.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s02-2.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s02-3.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s03.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s04.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vps-504/vps-504-s05.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vdk/vdk-200-a-s02.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/dungs-vdk/vdk-200-a-s02-211224.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/vpm-vc/vpm-vc-0.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/vpm-vc/vpm-vc-264114.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/vpm-vc/vpm-vc-258625.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/montazhnyj-nabor-dlja-vdk-i-klapanov-rp-1-1-2-rp-2.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/adapter-dlja-vps-i-klapanov-mv-zr-rp-1-1-2-rp-2.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/adapter-dlja-vps-i-klapanov-mv-zr-dn-40-dn-80.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/montazhnyj-nabor-dlja-vdk-i-klapanov-dn-40-dn-150.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/montazhnyj-nabor-dlja-vps-art-221503-dungs.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/filtr-vstavka-dlja-vps-art-243801-dungs.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/shteker-dlja-vps-art-231807-dungs190912131134.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/tcokol-dlja-vpm-vc-258621-dungs.html',
           'https://dungs.rusmark.ru/bloki-kontrolja-germetichnosti-dungs/prinadlejnosti-vps/predohranitel-dungs-231780.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-5/gf-505-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-5/gf-507-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-5/gf-510-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-5/gf-515-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-5/gf-520-1-0.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40040-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40050-4-0.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40065-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40080-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40100-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40125-0.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40150-0.html',
           'https://dungs.rusmark.ru/fil-try-dungs/tip-gf-40/gf-40200-0.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-505-507-gf-507-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-510-515-gf-515-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-520-1-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40040-3.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40050-3.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40050-4-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40065-3-gf-40065-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40080-3-gf-40080-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40100-3-gf-40100-4.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40125-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40150-1.html',
           'https://dungs.rusmark.ru/fil-try-dungs/fil-trujuschie-vstavki-dlja-gf/gf-40200-1.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-30-p-230-03-0.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-30-q-230-10-0.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-30-a-230-10-0.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-40-p-230-02-3.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-30-q-230-10-3.html',
           'https://dungs.rusmark.ru/privody-dungs/privody-tip-dma/dma-30-a-230-10-3.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-507-15-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-507-17-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-510-17-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-510-21-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-512-17-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-512-21-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-512-25-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-515-26-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-515-32-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-520-40-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-rp3-4-2/dmk-520-46-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5040-40-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5050-50-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5065-65-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5080-80-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5100-100-mm.html',
           'https://dungs.rusmark.ru/zaslonki-dungs/dmk-dn-40-125/dmk-5125-125-mm.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5002-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5003-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5005-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5007-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5010-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5012-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5015-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ia/kn-5020-ia.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5002-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5003-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5005-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5007-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5010-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5012-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5015-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-50-ii/kn-5020-ii.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160025.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160032.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160040.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160050.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160065.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160080.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160100.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160125.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160150.html',
           'https://dungs.rusmark.ru/krany-dungs/tip-kn-160/kn-160200.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/infrakrasnyi-datchik-plameni-dungs-flw-10-ir-255216.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/infrakrasnyi-datchik-plameni-dungs-flw-20-uv-250733.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/modul-rshireniya-em1-x-258672.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/zhidkokristalicheskiy-ekran-dungs-am25-v0-1-264172.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/topochnyi-avtomat-mpa4112-v1-1-259088.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/menedjer-goreniya-dungs-mpa22-s02-250780.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/modul-analogovyi-dungs-w-fm-em-3-3-264064.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/topochnyi-avtomat-mpa4112-v1-1-259066.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/dez-100-transformator-goreniya-252113.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/soedenitelniy-kabel-dez-252119.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/kabel-zajiganiya-dez-252121.html',
           'https://dungs.rusmark.ru/avtomatika-dungs/dez-200-transformator-goreniya-252114.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-20.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-30.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-50.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-100.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-120.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-150.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-200.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-250.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-280.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-300.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-400.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-500.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-60e.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-61e.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-70e.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-1011-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-1211-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-1212-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-1411-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/elektromagnitnaya-katushka/magnet-nr-1511-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/diodnyj-most-209922.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/diodnyj-most-211866.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/elektrovyprjamitel-211867.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/elektrovyprjamitel-211872f.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/elektrovyprjamitel-225258.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/diodnyj-most-222939.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/elektrovyprjamitel-222095c.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/elektrovyprjamitel-211869.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263235.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263234.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263236.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263237.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-1210-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263239.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263240.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-263241.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-1201-magnet-nr-1250.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-1231-magnet-nr-1350.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-1105.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/plata-upravlenija/magnet-nr-1111-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/gidrotormoz/h10.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/gidrotormoz/h11.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/gidrotormoz/h12-3-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/gidrotormoz/h12-6-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/gidrotormoz/h13-1.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-217471.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-217472.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-222341-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-222342-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-240506-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-222344-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-221884-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-221926-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-215384-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-241953-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-231230-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-231231-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-241956-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-231232-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-225528-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-225532-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-s-nippelem-225531-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-zapal-noj-linii-219006.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanec-dungs-219007.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/flanec/flanets-134300.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mbc-300.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mbc-700.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mbc-1200.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mb-403-053.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mb-405-407.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mb-410-412.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/fil-trujuschaja-vstavka/mb-415-420.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-515-frs-5040.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-515-40-frng-5040.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-520-frs-5050.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-520-frng-5050.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-5065.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-5065.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-5080.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-5080.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-5100.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-5100.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-5125.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-5125.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frs-5150.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/remkomplekty-frs-frng/remkomplekt-dlja-frng-5150.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-5080-8.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-1.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-2.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-3.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-4.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-5.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-6.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-7.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-503-8.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-1.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-2.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-3.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-5.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-6.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-7.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-505-8.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-507-0.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/pruzhiny-dungs/pruzhina-dlja-507-1.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/ukazatel-polozhenija/ukazatel-polozhenija-zakryto-k01-1.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/shteker-elektropodkljuchenija/shteker-dlja-klapana-4-h-kontaktnyj-chernyj.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/shteker-elektropodkljuchenija/shteker-dlja-vps-art-231807-dungs190912131134.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/shteker-elektropodkljuchenija/shteker-dungs-210318.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/adapter-dlja-vps-i-klapanov-mv-zr-rp-1-1-2-rp-2.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/adapter-dlja-vps-i-klapanov-mv-zr-dn-40-dn-80.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-vdk-i-klapanov-rp-1-1-2-rp-2.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-vdk-i-klapanov-dn-40-dn-150.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-vps-art-221503-dungs.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/filtr-vstavka-dlja-vps-art-243801-dungs.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/nabor-dlya-sborki-ldjynogo-rele-dungs-213910.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/ugol-krepezhny-dungs-230274.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnaya-plata-dungs-230302.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/shtucer-dungs-230306.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/shlang-dungs-1203405.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/oporniy-ugolok-metallicheskiy-dungs-230288.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/komplekt-dlya-montazha-dungs-242771.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/komplekt-dlya-montazha-dungs-223280.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-fri-224093.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-fri-dmv-219967.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-fri-dmv-219968.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/montazhnyi-nabor/montazhnyj-nabor-dlja-fri-224094.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/nippel/izmeritelnyi-patrubok-dungs-1-8-230397190807134508.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/nippel/izmeritelnyi-patrubok-dungs190808115350.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/nippel/nippel-privarnoy-dungs-1203402.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/perehodnik/perehodnik-dungs-1203406.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/perehodnik/perehodnik-dungs-1500395.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/zaglushka-g-1-4/zaglushka-3-4-dungs-1203404.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/lampa/nabor-dlya-montazha-lamp-231772.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/lampa/nabor-dlya-montazha-lamp-231773.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/lampa/nabor-dlya-montazha-lamp-231774.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/lampa/nabor-dlya-montazha-lamp-248239.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/lampa/nabor-dlya-montazha-lamp-248240.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/predohranitel/predohranitel-dungs-231780.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/tsokol/tcokol-dlja-vpm-vc-258621-dungs.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/uplotnitel-noe-kol-tso-dlja-flantsa-242117.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/uplotnitel-noe-kol-tso-dlja-flantsa-242118.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/uplotnitel-noe-kol-tso-dlja-flantsa-242119.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/uplotnitel-noe-kol-tso-dlja-flantsa-230444.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/shestigrannoe-kol-tso-231574.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/shestigrannoe-kol-tso-230442.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/uplotnitelnoe-koltso/shestigrannoe-kol-tso-230443.html',
           'https://dungs.rusmark.ru/zapchasti-dungs/potentiometre/potentiometre-dungs-240498.html']
    u = 0
    fieldnames = ['product_sku', 'category_path', 'product_name', 'slug', 'product_s_desc', 'product_desc',
                  'file_url', 'file_meta', 'custom_title', 'custom_value', 'product_in_stock', 'product_price',
                  'manufacturer_name', 'customtitle', 'metadesc', 'metakey', 'published']
    defin = []
    ak = []
    lol = []
    x = 0
    src = 'pdf'

    for url in ref:
        data = []
        y = ''
        r = requests.get(url)
        soup = bs4.BeautifulSoup(r.text, 'html.parser')
        shor = soup.find('div', id='short_description').findAll('li')
        r = re.compile("^[а-яА-Я-]+$")
        head = str(soup.find('div', class_='product_header').find('h1').text).split()
        try:
            docs = soup.find('div', class_='right-part').findAll('div')
        except AttributeError:
            docs = None
        try:
            russian = ' '.join([w for w in filter(r.match, head)])
            russian = russian + ' '
        except ValueError:
            russian = str([w for w in filter(r.match, head)]) + ' ' + 'DUNGS'
        category = str(soup.find('a', class_='pathway').find('span').text)
        category1 = ''
        try:
            if category == 'Датчики DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'дифференциальное реле' or 'датчик' in head:
                    category1 = 'Реле давления'
                else:
                    category1 = 'Аксесуары и принадлежности'
            elif category == 'Регуляторы DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'давления' in head:
                    category1 = 'Регуляторы давления'
                elif 'газ' in head:
                    category1 = 'Газовые клапаны'
                else:
                    category1 = 'Аксесуары и принадлежности'
            elif category == 'Клапаны DUNGS':
                category1 = 'Газовые клапаны'
            elif category == 'Мультиблоки DUNGS':
                category1 = 'Газовые мультиблоки'
            elif category == 'Блоки контроля герметичности DUNGS':
                category1 = 'Блоки контроля герметичности'
            elif category == 'Фильтры DUNGS':
                category1 = 'Аксесуары и принадлежности'
            elif category == 'Приводы DUNGS':
                category1 = 'Сервоприводы для воздушных заслонок'
            elif category == 'Заслонки DUNGS':
                category1 = 'Сервоприводы для воздушных заслонок'
            elif category == 'Краны DUNGS':
                category1 = 'Аксесуары и принадлежности'
            elif category == 'Автоматика DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'датчик' in head:
                    category1 = 'Датчики пламени'
                elif 'автомат' or 'горения' in head:
                    category1 = 'Топочные автоматы'
                elif 'поджига' in head:
                    category1 = 'Трансформаторы поджига'
                else:
                    category1 = 'Электрокомпоненты'
            elif category == 'Запчасти DUNGS':
                category1 = 'Аксесуары и принадлежности'
        except AttributeError:
            x += 1
            print('фееее')
        m = ''.join(shor[0].text.split()[1:])
        k = m + '~' + russian.replace(' фирмы', '').capitalize() + 'DUNGS' + '~' + 'Dungs' + '~' + category1
        shor.pop(0)
        shor.pop(0)
        s = ''
        if shor is not None:
            for h in shor:
                h = h.text
                h = h.split(':')
                h.pop(0)
                h = ''.join(h)
                k += '~' + h[1:]
                s += '~' + h[1:]
        c = ''
        try:
            for d in docs:
                y = src + str(d.findAll('a')[0].get_attribute_list("href")).replace(
                    "['https://dungs.rusmark.ru/components/com_jshopping/files/demo_products", '').replace("']", '')
                d = ' '.join(d.findAll('a')[1].text.split()) + '.pdf'
                c += '~' + d + '|pdf/' + y
        except AttributeError:
            c = ''
        k = k + s + c
        print(k)
        lol.append(k)

        """try:
            price = math.ceil(int(soup.find('span', id='block_price').find('span').text[:-3]) / 100) * 100
            if price == 0:
                price = str(price)
                price = 'По запросу'
            else:
                price = str("{0:.2f}".format(float(price)))
        except TypeError:
            price = 'По запросу'
        lol.append(price)
        print(price)
        category = str(soup.find('a', class_='pathway').find('span').text)
        category1 = ''
        try:
            if category == 'Датчики DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'дифференциальное реле' or 'датчик' in head:
                    category1 = 'Реле давления'
                else:
                    category1 = 'Аксесуары и принадлежности'
            elif category == 'Регуляторы DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'давления' in head:
                    category1 = 'Регуляторы давления'
                elif 'газ' in head:
                    category1 = 'Газовые клапаны'
                else:
                    category1 = 'Аксесуары и принадлежности'
            elif category == 'Клапаны DUNGS':
                category1 = 'Газовые клапаны'
            elif category == 'Мультиблоки DUNGS':
                category1 = 'Газовые мультибои'
            elif category == 'Блоки контроля герметичности DUNGS':
                category1 = 'Блоки контроля герметичности'
            elif category == 'Фильтры DUNGS':
                category1 = 'Аксесуары и принадлежности'
            elif category == 'Приводы DUNGS':
                category1 = 'Сервоприводы для воздушных заслонок'
            elif category == 'Заслонки DUNGS':
                category1 = 'Сервоприводы для воздушных заслонок'
            elif category == 'Краны DUNGS':
                category1 = 'Аксесуары и принадлежности'
            elif category == 'Автоматика DUNGS':
                head = str(soup.find('div', class_='product_header').find('h1').text)
                if 'датчик' in head:
                    category1 = 'Датчики пламени'
                elif 'автомат' or 'горения' in head:
                    category1 = 'Топочные автоматы'
                elif 'поджига' in head:
                    category1 = 'Трансформаторы поджига'
                else:
                    category1 = 'Электрокомпоненты'
            elif category == 'Запчасти DUNGS':
                category1 = 'Аксесуары и принадлежности'
        except AttributeError:
            x += 1
            print('фееее')
        category = category1 + '/' + category
        print(category)
        ak.append(category)"""
    csv_writer1(lol)


if __name__ == '__main__':
    output()
